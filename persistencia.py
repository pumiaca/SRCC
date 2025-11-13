import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def leer(archivo:str):
    """
    Leer datos de un archivo JSON.
    Atributos:
    - archivo: nombre del archivo sin extensión .json
    Retorna:
        Si el archivo existe, devuelve una lista de diccionarios con los datos del archivo.
        Si el archivo no existe, devuelve una lista vacía.
    """
    datos = []
    try:
        with open(f"./src/{archivo}.json", "r", encoding="utf-8") as archivo:
            datos = json.load(archivo)
        return datos
    except FileNotFoundError:
        return []

def cargar(archivo:str, dato:dict):
    """
    Cargar 1 registro de un archivo JSON.
    Atributos:
        - archivo: nombre del archivo sin extensión .json
        - dato: diccionario con el registro a cargar
    Retorna:
        Si el archivo existe, devuelve una lista de diccionarios con los datos del archivo.
        Si el archivo no existe, devuelve el error.
    """
    datos = leer(archivo)
    datos.append(dato)
    try:
        with open(f"./src/{archivo}.json", "w", encoding="utf-8") as archivo:
            json.dump(datos, archivo)
        return datos
    except Exception as e:
        return f"Error de Carga: {e}"

def actualizar(archivo:str, dato:dict):
    """
    Actualizar 1 registro de un archivo JSON.
    Atributos:
        - archivo: nombre del archivo sin extensión .json
        - dato: diccionario con el registro a actualizar
    Retorna:
        Si el archivo existe, devuelve una lista de diccionarios con los datos del archivo.
        Si el archivo no existe, devuelve el error.
    
    """
    datos = leer(archivo)
    nuevos_datos = []
    for item in datos:
        if(item["id"] != dato["id"]):
            nuevos_datos.append(item)
        else:
            nuevos_datos.append(dato)
    
    try:
        with open(f"./src/{archivo}.json", "w", encoding="utf-8") as archivo:
            json.dump(nuevos_datos, archivo)
        return nuevos_datos
    except Exception as e:
        return f"Error de Carga: {e}"

def borrar(archivo:str, dato:dict):
    """
    Borrar 1 registro de un archivo JSON.
    Atributos:
        - archivo: nombre del archivo sin extensión .json
        - dato: diccionario con el registro a borrar
    Retorna:
        Si el archivo existe, devuelve una lista de diccionarios con los datos del archivo.
        Si el archivo no existe, devuelve el error.
    """
    datos = leer(archivo)
    datos = [item for item in datos if item["id"] != dato["id"]]
    try:
        with open(f"./src/{archivo}.json", "w", encoding="utf-8") as archivo:
            json.dump(datos, archivo)
        return datos
    except Exception as e:
        return f"Error de Carga: {e}"

def buscar_id(archivo:str, dato:dict):
    """
    Buscar 1 registro de un archivo JSON.
    Atributos:
        - archivo: nombre del archivo sin extensión .json
        - dato: diccionario con el registro a buscar
    Retorna:
        Si el archivo existe, devuelve el primer diccionario que coincida con el id del dato buscado.
        Si el archivo no existe, devuelve el error.
    """
    datos = leer(archivo)
    registro = list(filter(lambda x: x["id"] == dato["id"], datos))
    return registro


def generar_reporte(archivo:str):
    """
    Exportar datos a un archivo Excel.
    Atributos:
        - archivo: nombre del archivo sin extensión .xlsx
        - datos: lista de diccionarios con los datos a exportar
    Retorna:
        Si el archivo se crea correctamente, devuelve True.
        Si ocurre un error, devuelve el error.
    """
    try:
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Reporte de Stock"
        # Encabezados
        datos = leer(archivo)
        headers = list(datos[0].keys())
        hoja.append(headers)

        for col_idx in range(1, 10):  # columnas 1 a 6 (A-F)
            col = get_column_letter(col_idx)
            celda = hoja[f'{col}1']
            celda.font = Font(bold=True, size=12)
            celda.alignment = Alignment(horizontal="center")

        tabla = [[p[h] for h in headers] for p in datos]
        for row in tabla:
            hoja.append(row)

        libro.save(f"./src/{archivo}.xlsx")
        return True
    except Exception as e:
        print(f"Error al exportar a Excel: {e}")
        return False
    

def limpiar_datos(archivo:str):
    """
    Limpia los datos de un archivo JSON.
    Atributos:
        - archivo: nombre del archivo sin extensión .json
    Retorna:
        Si el archivo se limpia correctamente, devuelve True.
        Si ocurre un error, devuelve el error.
    """
    try:
        with open(f"./src/{archivo}.json", "w", encoding="utf-8") as archivo:
            json.dump([], archivo)
        return True
    except Exception as e:
        return f"Error al limpiar datos: {e}"